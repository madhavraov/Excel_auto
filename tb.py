from openpyxl import load_workbook
import os
import re


class Excel:

    def __init__(self, file, source_folder, target_folder):
        self.get_sheet = None
        self.data_to_file = None
        self.file_lists = None
        self.sheet_data = None
        self.data = None
        self.source_sheet = None
        self.file = file
        self.source_folder = source_folder
        self.target_folder = target_folder

    def get_destination_folder(self):
        extension = '.xlsx'
        self.file_lists = []
        for file in os.listdir(self.source_folder):
            if file.endswith(extension):
                self.file_lists.append(file)
        if len(self.file_lists) < 1:
            raise FileNotFoundError

    def load_file(self):
        try:
            self.data = load_workbook(self.file)
            self.source_sheet = self.data['Sheet1']
        except PermissionError:
            raise PermissionError
        except FileNotFoundError:
            raise FileNotFoundError

    def save_data(self):
        for files in self.file_lists:
            destination_file = load_workbook(filename=f'{self.source_folder}/{files}')
            destination_sheet = destination_file['GL']
            destination_sheet.delete_rows(1, destination_sheet.max_row + 1)
            for i in range(1,self.source_sheet.max_row + 1):
                for j in range(1, self.source_sheet.max_column + 1):
                    if "=" in str(self.source_sheet.cell(row=i, column=j).value):
                        self.source_sheet.cell(row=i, column=j).value = re.sub('^=', ' =', self.source_sheet.cell(row=i, column=j).value)
                    destination_sheet.cell(row=i, column=j).value = self.source_sheet.cell(row=i, column=j).value
            destination_file.save(f'{self.target_folder}/{files}')
