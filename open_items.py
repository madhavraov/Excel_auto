import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill



class OpenItems:

    def __init__(self, file_path, source_path, target_path):
        self.bsr_lists = None
        self.gl_account = None
        self.source_file = None
        self.file_path = file_path
        self.source_path = source_path
        self.output_path = target_path

    def get_destination_folder(self):
        extension = '.xlsx'
        self.bsr_lists = []
        for file in os.listdir(self.source_path):
            if file.endswith(extension):
                self.bsr_lists.append(file)
        if len(self.bsr_lists) < 1:
            raise FileNotFoundError

    def get_source_file(self):
        try:
            self.source_file = pd.read_excel(self.file_path)
            self.gl_account = self.source_file['GLAccount'].unique().astype(str)
        except FileNotFoundError:
            raise FileNotFoundError

    def data_to_destination(self):
        for bsr in self.bsr_lists:
            file = pd.ExcelFile(f'{self.source_path}/{bsr}')
            sheet_names = file.sheet_names
            for gl in self.gl_account:
                for sheet in sheet_names:
                    if gl == sheet:
                        data = self.source_file.query(f'GLAccount == {gl}')
                        with pd.ExcelWriter(path=f'{self.source_path}/{bsr}', mode='a', engine='openpyxl',
                                            if_sheet_exists='replace') as writer:
                            data.to_excel(writer, sheet_name=gl, index=False, startrow=0, header=True)
                            #red_font = Font(color='000000', bold=False)
                            wb = load_workbook(f'{self.source_path}/{bsr}')
                            ws = wb[f'{sheet}']
                            for cell in ws["1:1"]:
                                cell.font = Font(color='000000', bold=False)
                                cell.fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type = "solid")
                                rd = ws.row_dimensions[1] # get dimension for row 1
                                rd.height = 39.5
                            wb.save(f'{self.output_path}/{bsr}')


